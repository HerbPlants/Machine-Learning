import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def load_to_excel(df, output_path):
    try:
        if os.path.exists(output_path):
            os.remove(output_path)

        # Pisahkan kolom input dan hasil scraping
        cols_input = ['Nama Tanaman', 'Nama Latin', 'Link']
        cols_output = ['Penyebaran Tanaman', 'Nama Lokal', 'Agroekologi', 'Khasiat']

        df_input = df[cols_input]
        df_output = df[cols_output]

        # Simpan input dulu
        df_input.to_excel(output_path, index=False)

        # Tambahkan hasil scraping ke kolom mulai dari kolom D (kolom ke-4, index 3) dan baris ke-2
        wb = load_workbook(output_path)
        ws = wb.active

        for row_idx, row in enumerate(dataframe_to_rows(df_output, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=4):  # mulai dari kolom D (index ke-4)
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(output_path)

    except PermissionError:
        print(f"Gagal menyimpan {output_path}: file sedang terbuka. Tutup dulu file Excel-nya lalu jalankan ulang.")